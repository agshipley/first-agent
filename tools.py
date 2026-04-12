import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import date

COLUMN_WIDTHS = {
    "A": 25,    # Company Name
    "B": 18,    # Type
    "C": 28,    # Location
    "D": 24,    # Geographic Area
    "E": 50,    # Why They're a Lead
    "F": 30,    # Company Website
    "G": 35,    # Source URL
    "H": 30,    # Potential Contact
    "I": 12,    # ICP Score
    "J": 20,    # Estimated Budget
    "K": 45,    # Budget Basis
    "L": 18,    # Budget Confidence
    "M": 22,    # Project Stage
    "N": 55,    # Notes
    "O": 14,    # Date Found
    "P": 35,    # Lead Source
}

HEADERS = [
    "Company Name",
    "Type",
    "Location",
    "Geographic Area",
    "Why They're a Lead",
    "Company Website",
    "Source URL",
    "Potential Contact",
    "ICP Score",
    "Estimated Budget",
    "Budget Basis",
    "Budget Confidence",
    "Project Stage",
    "Notes",
    "Date Found",
    "Lead Source",
]

TABLE_NAMES = {
    "Corporate": "CorporateLeads",
    "Public Sector": "PublicSectorLeads"
}

def _migrate_schema_if_needed(sheet) -> bool:
    """
    Runs all pending schema migrations in order. Returns True if any migration ran.

    v1 → v2: insert Geographic Area at column D
      Detection: col D header != "Geographic Area"

    v2 → v3: insert Estimated Budget, Budget Basis, Budget Confidence at column J
      Detection: col J header == "Notes" (i.e. budget cols not yet present)

    v3 → v4: insert Project Stage at column M
      Detection: col M header == "Notes" (i.e. project stage not yet present)
    """
    if sheet.max_row < 1:
        return False

    migrated = False

    # Migration 1: Add Geographic Area column (v1 → v2)
    if sheet.cell(row=1, column=4).value != "Geographic Area":
        sheet.insert_cols(4)
        sheet.cell(row=1, column=4).value = "Geographic Area"
        migrated = True

    # Migration 2: Add budget columns (v2 → v3)
    # After migration 1, Notes sits at column J (index 10)
    if sheet.cell(row=1, column=10).value == "Notes":
        sheet.insert_cols(10, amount=3)
        sheet.cell(row=1, column=10).value = "Estimated Budget"
        sheet.cell(row=1, column=11).value = "Budget Basis"
        sheet.cell(row=1, column=12).value = "Budget Confidence"
        migrated = True

    # Migration 3: Add Project Stage column (v3 → v4)
    # After migrations 1+2, Notes sits at column M (index 13)
    if sheet.cell(row=1, column=13).value == "Notes":
        sheet.insert_cols(13)
        sheet.cell(row=1, column=13).value = "Project Stage"
        migrated = True

    return migrated

def get_existing_company_names(sheet) -> set:
    existing_names = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_names.add(row[0].strip().lower())
    return existing_names

def get_existing_leads_for_segment(segment: str) -> list[str]:
    DATA_DIR = os.environ.get("DATA_DIR", ".")
    filename = os.path.join(DATA_DIR, "leads.xlsx")
    sheet_names = {"corporate": "Corporate", "public_sector": "Public Sector"}
    sheet_name = sheet_names.get(segment, "Corporate")

    if not os.path.exists(filename):
        return []

    workbook = openpyxl.load_workbook(filename)
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    names = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            names.append(row[0].strip())
    return names

def get_all_leads_for_segment(segment: str) -> list[dict]:
    DATA_DIR = os.environ.get("DATA_DIR", ".")
    filename = os.path.join(DATA_DIR, "leads.xlsx")
    sheet_names = {"corporate": "Corporate", "public_sector": "Public Sector"}
    sheet_name = sheet_names.get(segment, "Corporate")

    if not os.path.exists(filename):
        return []

    workbook = openpyxl.load_workbook(filename)
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    leads = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            leads.append({
                "company_name": row[0] or "",
                "type": row[1] or "",
                "location": row[2] or "",
                "geographic_area": row[3] or "",
                "why_a_lead": row[4] or "",
                "company_website": row[5] or "",
                "source_url": row[6] or "",
                "potential_contact": row[7] or "",
                "icp_score": row[8] or 0,
                "estimated_budget": row[9] or "",
                "budget_basis": row[10] or "",
                "budget_confidence": row[11] or "",
                "project_stage": row[12] or "" if len(row) > 12 else "",
                "notes": row[13] or "" if len(row) > 13 else "",
                "date_found": str(row[14]) if len(row) > 14 and row[14] else "",
                "lead_source": row[15] or "" if len(row) > 15 else "",
            })
    return leads

DEEP_DIVE_HEADERS = [
    "Project Status",
    "News Summary",
    "Existing Art Attachments",
    "Key Principals",
    "Commissioning History",
    "Deep Dive Date",
]

DEEP_DIVE_COLUMN_WIDTHS = {
    "Q": 60,  # Project Status
    "R": 60,  # News Summary
    "S": 60,  # Existing Art Attachments
    "T": 60,  # Key Principals
    "U": 60,  # Commissioning History
    "V": 14,  # Deep Dive Date
}

def _apply_table_formatting(sheet, sheet_name, max_col=None):
    sheet.tables.clear()

    max_row = sheet.max_row
    if max_row < 2:
        return

    if max_col is None:
        max_col = len(HEADERS)
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table_name = TABLE_NAMES.get(sheet_name, "Leads")

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    sheet.add_table(table)

    for col_letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[col_letter].width = width

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = wrap_alignment

def save_leads_to_spreadsheet(leads: list[dict], segment: str = "corporate") -> tuple[str, list[dict]]:
    """
    Returns a tuple: (result_message, actually_saved_leads)
    """
    DATA_DIR = os.environ.get("DATA_DIR", ".")
    filename = os.path.join(DATA_DIR, "leads.xlsx")
    sheet_names = {"corporate": "Corporate", "public_sector": "Public Sector"}
    sheet_name = sheet_names.get(segment, "Corporate")

    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.active.title = "Corporate"

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        _migrate_schema_if_needed(sheet)
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(HEADERS)

    if sheet.cell(row=1, column=1).value != HEADERS[0]:
        sheet.insert_rows(1)
        for col_idx, header in enumerate(HEADERS, 1):
            sheet.cell(row=1, column=col_idx, value=header)

    existing_names = get_existing_company_names(sheet)

    saved_count = 0
    skipped_count = 0
    actually_saved = []
    today = date.today().strftime("%Y-%m-%d")

    for lead in leads:
        company_name = lead.get("company_name", "").strip()
        if company_name.lower() in existing_names:
            skipped_count += 1
            print(f"Skipping duplicate: {company_name}")
        else:
            row = [
                company_name,
                lead.get("type", ""),
                lead.get("location", ""),
                lead.get("geographic_area", ""),
                lead.get("why_a_lead", ""),
                lead.get("company_website", ""),
                lead.get("source_url", ""),
                lead.get("potential_contact", ""),
                lead.get("icp_score", ""),
                lead.get("estimated_budget", ""),
                lead.get("budget_basis", ""),
                lead.get("budget_confidence", ""),
                lead.get("project_stage", ""),
                lead.get("notes", ""),
                today,
                lead.get("lead_source", "Web Search"),
            ]
            sheet.append(row)
            saved_count += 1
            existing_names.add(company_name.lower())
            actually_saved.append(lead)

    _apply_table_formatting(sheet, sheet_name)
    workbook.save(filename)
    message = f"Saved {saved_count} new leads to '{sheet_name}', skipped {skipped_count} duplicates."
    return message, actually_saved


def save_deep_dive_to_spreadsheet(report: dict) -> str:
    """
    Finds the lead row matching report['company_name'] across all sheets,
    appends deep dive column headers if not already present, and writes
    the section findings into that row. Returns a status message.
    """
    DATA_DIR = os.environ.get("DATA_DIR", ".")
    filename = os.path.join(DATA_DIR, "leads.xlsx")

    if not os.path.exists(filename):
        return "leads.xlsx not found"

    workbook = openpyxl.load_workbook(filename)
    company_name_lower = report.get("company_name", "").strip().lower()
    sections = report.get("report_sections", {})
    today = date.today().strftime("%Y-%m-%d")

    # Search all sheets for the matching row
    target_sheet = None
    target_row_idx = None
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_idx in range(2, sheet.max_row + 1):
            cell_val = sheet.cell(row=row_idx, column=1).value
            if cell_val and cell_val.strip().lower() == company_name_lower:
                target_sheet = sheet
                target_row_idx = row_idx
                break
        if target_sheet:
            break

    if not target_sheet:
        return f"Lead '{report.get('company_name')}' not found in spreadsheet"

    # Build header map: {header_name: 1-based column index}
    header_map = {}
    for col_idx in range(1, target_sheet.max_column + 1):
        val = target_sheet.cell(row=1, column=col_idx).value
        if val:
            header_map[val] = col_idx

    # Append any missing deep dive headers after the last occupied column
    next_col = max(header_map.values()) + 1 if header_map else 1
    for header in DEEP_DIVE_HEADERS:
        if header not in header_map:
            target_sheet.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
            next_col += 1

    # Write findings into the target row
    section_map = {
        "Project Status":           sections.get("project_status", {}).get("findings", ""),
        "News Summary":             sections.get("news_and_media", {}).get("findings", ""),
        "Existing Art Attachments": sections.get("existing_art_attachments", {}).get("findings", ""),
        "Key Principals":           sections.get("key_principals", {}).get("findings", ""),
        "Commissioning History":    sections.get("commissioning_history", {}).get("findings", ""),
        "Deep Dive Date":           today,
    }
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for header, value in section_map.items():
        col = header_map.get(header)
        if col:
            cell = target_sheet.cell(row=target_row_idx, column=col, value=value)
            cell.alignment = wrap_alignment

    # Apply column widths for deep dive columns
    for col_letter, width in DEEP_DIVE_COLUMN_WIDTHS.items():
        target_sheet.column_dimensions[col_letter].width = width

    # Reformat table to cover all columns including the new ones
    full_max_col = max(header_map.values())
    _apply_table_formatting(target_sheet, target_sheet.title, max_col=full_max_col)

    workbook.save(filename)
    return f"Deep dive data saved for '{report.get('company_name')}'"
