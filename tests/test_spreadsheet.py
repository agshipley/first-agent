"""
Tests for tools.py — spreadsheet read/write operations.
All writes go to a temporary directory; leads.xlsx is never touched.
"""

import os
import pytest
import openpyxl
from datetime import date

from tools import save_leads_to_spreadsheet, HEADERS


FULL_LEAD = {
    "company_name": "Acme Development Group",
    "type": "Real Estate Developer",
    "location": "Los Angeles, CA",
    "geographic_area": "Greater Los Angeles Area",
    "why_a_lead": "Mixed-use development with $50M valuation in DTLA",
    "company_website": "https://acmedev.example.com",
    "source_url": "https://source.example.com/permit/12345",
    "potential_contact": "Jane Smith, VP Development",
    "icp_score": 8,
    "estimated_budget": "$500K–$750K",
    "budget_basis": "1% of $50M construction cost per PADFP",
    "budget_confidence": "High",
    "project_stage": "Pre-Construction",
    "notes": "New mixed-use tower, 40 floors",
    "lead_source": "LA Permits Monitor",
}


class TestSaveLeadsToSpreadsheet:

    def test_all_headers_present(self, tmp_data_dir):
        save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        wb = openpyxl.load_workbook(tmp_data_dir / "leads.xlsx")
        sheet = wb["Corporate"]
        actual_headers = [sheet.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
        assert actual_headers == HEADERS

    def _first_data_row(self, sheet):
        """Return the first row number that has a non-None value in column 1.
        openpyxl's Workbook() leaves a phantom empty row 2 on fresh files,
        so data may start at row 3 rather than row 2.
        """
        for row_num in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_num, column=1).value is not None:
                return row_num
        raise AssertionError("No data rows found in sheet")

    def test_lead_written_to_correct_columns(self, tmp_data_dir):
        save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        wb = openpyxl.load_workbook(tmp_data_dir / "leads.xlsx")
        sheet = wb["Corporate"]
        data_row = self._first_data_row(sheet)
        row = [sheet.cell(row=data_row, column=i).value for i in range(1, len(HEADERS) + 1)]

        assert row[0] == "Acme Development Group"         # Company Name
        assert row[1] == "Real Estate Developer"          # Type
        assert row[2] == "Los Angeles, CA"                # Location
        assert row[3] == "Greater Los Angeles Area"       # Geographic Area
        assert row[8] == 8                                # ICP Score
        assert row[9] == "$500K–$750K"                    # Estimated Budget
        assert row[10] == "1% of $50M construction cost per PADFP"   # Budget Basis
        assert row[11] == "High"                          # Budget Confidence
        assert row[12] == "Pre-Construction"              # Project Stage
        assert row[15] == "LA Permits Monitor"            # Lead Source

    def test_date_found_populated_automatically(self, tmp_data_dir):
        save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        wb = openpyxl.load_workbook(tmp_data_dir / "leads.xlsx")
        sheet = wb["Corporate"]
        data_row = self._first_data_row(sheet)
        date_val = sheet.cell(row=data_row, column=15).value   # column O = Date Found
        today = date.today().strftime("%Y-%m-%d")
        assert date_val == today

    def test_missing_optional_fields_does_not_crash(self, tmp_data_dir):
        minimal = {"company_name": "Minimal Corp"}
        msg, saved = save_leads_to_spreadsheet([minimal], "corporate")
        assert len(saved) == 1

    def test_multiple_leads_saved_as_multiple_rows(self, tmp_data_dir):
        leads = [
            {**FULL_LEAD, "company_name": f"Company {i}"} for i in range(5)
        ]
        save_leads_to_spreadsheet(leads, "corporate")
        wb = openpyxl.load_workbook(tmp_data_dir / "leads.xlsx")
        sheet = wb["Corporate"]
        # Count rows with actual data in column 1 (skip header row 1 and any phantom empty rows)
        data_rows = [
            r for r in sheet.iter_rows(min_row=2, values_only=True)
            if r[0] is not None
        ]
        assert len(data_rows) == 5

    def test_duplicate_company_skipped(self, tmp_data_dir):
        save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        msg, saved = save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        assert len(saved) == 0
        assert "skipped" in msg.lower()

    def test_public_sector_sheet_created(self, tmp_data_dir):
        lead = {**FULL_LEAD, "company_name": "City of LA Arts"}
        save_leads_to_spreadsheet([lead], "public_sector")
        wb = openpyxl.load_workbook(tmp_data_dir / "leads.xlsx")
        assert "Public Sector" in wb.sheetnames

    def test_returns_tuple_message_and_saved_list(self, tmp_data_dir):
        result = save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        assert isinstance(result, tuple)
        msg, saved = result
        assert isinstance(msg, str)
        assert isinstance(saved, list)
        assert len(saved) == 1

    def test_saved_lead_contains_company_name(self, tmp_data_dir):
        _, saved = save_leads_to_spreadsheet([FULL_LEAD], "corporate")
        assert saved[0]["company_name"] == "Acme Development Group"
